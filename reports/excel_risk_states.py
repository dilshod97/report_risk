from __future__ import annotations

import datetime as dt
from copy import copy
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook.properties import CalcProperties

import config
from db import run_query_rows

TEMPLATE_PATH = config.TEMPLATES_DIR / "risk_states_template.xlsx"
DATA_SHEET = "Рўйхат"

# Рўйхат tuzilishi:
#   1-qator   - "Жами" (=SUM(...G3:G9999) formulalari)
#   2-qator   - SQL ustun nomlari (risk_name, vaz_code, ...)
#   3-qatordan- ma'lumot
# center_staff query natijasi ustunlari A..AJ (1..36) ga pozitsion 1:1.
HEADER_ROW = 2
FIRST_DATA_ROW = 3
DATA_COLUMN_COUNT = 36  # A..AJ

# Har bir ma'lumot qatoriga qo'yiladigan yordamchi ustun formulalari (AK..AR).
# AK..AO - "1 yildan oshgan" (jami minus vaqt oraliqlari); AP..AR - pivotlarga
# tegishli klassifikatsiya (SUMIF baribir A/C/E ustunlar bo'yicha ishlaydi).
HELPER_FORMULAS = {
    37: "=G{r}-L{r}-Q{r}-V{r}-AA{r}-AF{r}",   # AK
    38: "=H{r}-M{r}-R{r}-W{r}-AB{r}-AG{r}",   # AL
    39: "=I{r}-N{r}-S{r}-X{r}-AC{r}-AH{r}",   # AM
    40: "=J{r}-O{r}-T{r}-Y{r}-AD{r}-AI{r}",   # AN
    41: "=K{r}-P{r}-U{r}-Z{r}-AE{r}-AJ{r}",   # AO
    42: "=+VLOOKUP(C{r},'Вазирликлар кесимида'!B:B,1,0)",  # AP
    43: "=VLOOKUP(A{r},'Риск йўналишлари'!B:AI,1,0)",       # AQ
    44: "=VLOOKUP(E{r},'Ходимлар кесимида'!B:B,1,0)",       # AR
}
LAST_HELPER_COLUMN = 44


def _cell_value(value):
    if isinstance(value, Decimal):
        return float(value)
    return value


def _apply_style(cell, style):
    if not style:
        return
    cell.font = copy(style["font"])
    cell.border = copy(style["border"])
    cell.fill = copy(style["fill"])
    cell.alignment = copy(style["alignment"])
    cell.number_format = style["number_format"]


def fill_workbook(columns: list[str], rows: list[tuple]):
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb[DATA_SHEET]

    # Birinchi data qatoridan uslub namunasini saqlab qolamiz.
    style_by_col = {}
    for col in range(1, LAST_HELPER_COLUMN + 1):
        src = ws.cell(row=FIRST_DATA_ROW, column=col)
        style_by_col[col] = {
            "font": copy(src.font),
            "border": copy(src.border),
            "fill": copy(src.fill),
            "alignment": copy(src.alignment),
            "number_format": src.number_format,
        }

    # Eski ma'lumot qatorlarini tozalash.
    if ws.max_row >= FIRST_DATA_ROW:
        ws.delete_rows(FIRST_DATA_ROW, ws.max_row - FIRST_DATA_ROW + 1)

    # Yangi ma'lumotni yozish.
    for i, row in enumerate(rows):
        r = FIRST_DATA_ROW + i
        for col in range(1, DATA_COLUMN_COUNT + 1):
            value = _cell_value(row[col - 1]) if col - 1 < len(row) else None
            cell = ws.cell(row=r, column=col, value=value)
            _apply_style(cell, style_by_col.get(col))
        for col, formula in HELPER_FORMULAS.items():
            cell = ws.cell(row=r, column=col, value=formula.format(r=r))
            _apply_style(cell, style_by_col.get(col))

    # Ochilganda formulalar (Жами, pivotlar, yordamchi ustunlar) qayta hisoblansin.
    wb.calculation = CalcProperties(calcId=0, fullCalcOnLoad=True)
    return wb


def build() -> Path:
    columns, rows = run_query_rows(config.QUERIES_DIR / "center_staff.sql")
    wb = fill_workbook(columns, rows)

    config.OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = config.OUTPUT_DIR / f"risk_holatlari_{dt.date.today().isoformat()}.xlsx"
    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    print(build())
