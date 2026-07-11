from __future__ import annotations

import datetime as dt
from copy import copy
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook.properties import CalcProperties

import config
from db import run_query_rows

TEMPLATE_PATH = config.TEMPLATES_DIR / "risk_groups_template.xlsx"
DATA_SHEET = "рўйхат"

# рўйхат sheet tuzilishi:
#   1-3 qatorlar   - yordamchi/bo'sh
#   4-7 qatorlar   - ko'p bosqichli chiroyli sarlavha
#   8-qator        - "Жами" (=SUM(X9:X9999) formulalari)
#   9-qator        - SQL ustun nomlari (vaz_code, katta_otasi_nomi, ...)
#   10-qatordan    - ma'lumot
# Query natijasi ustunlari A..CP (1..94) ga pozitsion 1:1 mos keladi.
HEADER_ROW = 9
FIRST_DATA_ROW = 10
DATA_COLUMN_COUNT = 94  # A..CP

# Hisobot shakllangan sana ko'rsatiladigan katakchalar. Manba - рўйхат!B3,
# 3 pivot esa uni '=рўйхат!B3' formulasi bilan oladi. Formula qayta
# hisoblanmaydigan ko'ruvchilarda eskirmasligi uchun sanani har bir katakka
# bevosita (statik qiymat) yozamiz.
DATE_CELLS = {
    DATA_SHEET: "B3",
    "Йўналишлар кесимида ": "BF3",
    "Вазирликлар кесимида (22)": "BF3",
    "йўналиш гурухлари кесимида": "CH4",
}

# Har bir ma'lumot qatoriga qo'yiladigan yordamchi ustun formulalari (CQ..CV).
# {r} - qator raqami bilan almashtiriladi. Pivot sheetlar shu ustunlarga
# (CS, CU, CV) tayanib SUMIFS orqali hisoblaydi.
HELPER_FORMULAS = {
    95: "=VLOOKUP(C{r},'Йўналишлар кесимида '!B:C,1,0)",       # CQ
    96: "=VLOOKUP(B{r},'Вазирликлар кесимида (22)'!B:CF,1,0)",  # CR
    97: "=VLOOKUP(D{r},справочник!D:E,2,0)",                    # CS
    98: "=VLOOKUP(CS{r},'йўналиш гурухлари кесимида'!B:B,1,0)",  # CT
    99: '=IFERROR(CR{r},"Бошқалар")',                           # CU
    100: "=VLOOKUP(C{r},справочник!A:B,2,0)",                   # CV
}
LAST_HELPER_COLUMN = 100


def _cell_value(value):
    if isinstance(value, Decimal):
        return float(value)
    return value


def fill_workbook(columns: list[str], rows: list[tuple]):
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb[DATA_SHEET]

    # Eski ma'lumot qatorlari uchun uslub namunasini (birinchi data qatoridan)
    # saqlab qolamiz, keyin yangi qatorlarga qo'llash uchun.
    style_by_col = {}
    for col in range(1, LAST_HELPER_COLUMN + 1):
        src = ws.cell(row=FIRST_DATA_ROW, column=col)
        style_by_col[col] = {
            "font": copy(src.font),
            "border": copy(src.border),
            "fill": copy(src.fill),
            "alignment": copy(src.alignment),
            "number_format": src.number_format,
        }

    # Eski ma'lumot qatorlarini (10-qatordan oxirigacha) tozalash.
    if ws.max_row >= FIRST_DATA_ROW:
        ws.delete_rows(FIRST_DATA_ROW, ws.max_row - FIRST_DATA_ROW + 1)

    # Yangi ma'lumotni yozish.
    for i, row in enumerate(rows):
        r = FIRST_DATA_ROW + i
        for col in range(1, DATA_COLUMN_COUNT + 1):
            value = _cell_value(row[col - 1]) if col - 1 < len(row) else None
            cell = ws.cell(row=r, column=col, value=value)
            _apply_style(cell, style_by_col.get(col))
        for col, formula in HELPER_FORMULAS.items():
            cell = ws.cell(row=r, column=col, value=formula.format(r=r))
            _apply_style(cell, style_by_col.get(col))

    # Hisobot shakllangan sanani barcha ko'rinadigan katakchalarga statik
    # qiymat sifatida yozamiz (formula emas - har qanday ko'ruvchida ko'rinsin).
    today = dt.datetime.combine(dt.date.today(), dt.time())
    for sheet_name, coord in DATE_CELLS.items():
        cell = wb[sheet_name][coord]
        fmt = cell.number_format
        cell.value = today
        cell.number_format = fmt

    # Excel/LibreOffice faylni ochganda barcha formulalar (Жами, pivotlar,
    # yordamchi ustunlar) qayta hisoblanishi uchun.
    wb.calculation = CalcProperties(calcId=0, fullCalcOnLoad=True)
    return wb


def _apply_style(cell, style):
    if not style:
        return
    cell.font = copy(style["font"])
    cell.border = copy(style["border"])
    cell.fill = copy(style["fill"])
    cell.alignment = copy(style["alignment"])
    cell.number_format = style["number_format"]


def build() -> Path:
    columns, rows = run_query_rows(config.QUERIES_DIR / "daily_summary.sql")
    wb = fill_workbook(columns, rows)

    config.OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = config.OUTPUT_DIR / f"risk_guruhlari_{dt.date.today().isoformat()}.xlsx"
    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    print(build())
