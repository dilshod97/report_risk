from __future__ import annotations

import datetime as dt
from functools import lru_cache
from pathlib import Path

from jinja2 import Environment, FileSystemLoader
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

import config
from db import run_query

TEMPLATE_PATH = config.TEMPLATES_DIR / "risk_groups_template.xlsx"
GROUP_SHEET = "йўналиш гурухлари кесимида"
REF_SHEET = "справочник"

_BLUE = "linear-gradient(90deg,#497bf1 0%,#82a7f9 100%)"
_GREEN = "linear-gradient(90deg,#08d189 0%,#96e9cb 100%)"
# Kartalar tartibi 'йўналиш гурухлари кесимида' B10:B18 bilan bir xil.
# HTML dizaynidagi kabi 5- va 6- kartalar yashil, qolganlari ko'k.
_GREEN_POSITIONS = {4, 5}

# "Тасдиғини топган" - sheetdagi CG/CH formulasi bo'yicha: faqat
# "назоратдан ечилган" (palata/vazirlik/system) ning oldi_olindi va
# keyingi_nazorat qismlari. Har bir juftlik (son_ustuni, summa_ustuni).
APPROVED_COLUMN_PAIRS = [
    ("nazoratdan_echildi_palata_oldini_oldi", "nazoratdan_echildi_palata_oldini_oldi_sums"),
    ("nazoratdan_echildi_palata_keyin_naz", "nazoratdan_echildi_palata_keyin_naz_sums"),
    ("nazoratdan_echildi_vazirlik_oldini_oldi", "nazoratdan_echildi_vazirlik_oldini_oldi_sums"),
    ("nazoratdan_echildi_vazirlik_keyin_naz", "nazoratdan_echildi_vazirlik_keyin_naz_sums"),
    ("nazoratdan_echildi_system_oldini_oldi", "nazoratdan_echildi_system_oldini_oldi_sums"),
    ("nazoratdan_echildi_system_keyin_naz", "nazoratdan_echildi_system_keyin_naz_sums"),
]

# Har bir metrika: (son_ustuni, summa_ustuni). Summalar ko'rsatishda /1000.
COL_DETECTED = ("all_count", "all_sums")
COL_CONTROL = ("nazoratda", "nazoratda_sums")
COL_NOT_CONTROLLED = ("nazoratga_olinmagan", "nazoatga_olinmagan_sums")
COL_CLOSED = ("nazoratdan_echildi", "nazoratdan_echildi_sums")

SUM_SCALE = 1000.0


def _num(value) -> float:
    return float(value) if value is not None else 0.0


def fmt_count(value: float) -> str:
    return f"{round(value):,}".replace(",", " ")


def fmt_sum(value: float) -> str:
    """Summalar sheetdagidek 1000 ga bo'linib, 1 kasr bilan ko'rsatiladi."""
    return f"{value / SUM_SCALE:,.1f}".replace(",", " ")


def fmt_pct(part: float, total: float) -> str:
    if not total:
        return "0"
    return str(round(part / total * 100))


@lru_cache(maxsize=1)
def _group_config() -> tuple[dict[str, str], list[str]]:
    """Shablondan guruhlash konfiguratsiyasini o'qiydi.

    Qaytaradi: (g_name -> qisqa guruh nomi mapping, 9 ta karta yorlig'i).
    Bu 'йўналиш гурухлари кесимида' sheeti bilan bir xil mantiq: рўйхат!CS =
    VLOOKUP(g_name, справочник!D:E, 2) va kartalar B10:B18.
    """
    wb = load_workbook(TEMPLATE_PATH, read_only=True, data_only=False)
    try:
        ref = wb[REF_SHEET]
        g_to_group: dict[str, str] = {}
        for row in ref.iter_rows(min_row=5, min_col=4, max_col=5, values_only=True):
            key, short = row
            if key is not None and short is not None:
                g_to_group[str(key).strip()] = str(short).strip()

        ws = wb[GROUP_SHEET]
        labels = [ws.cell(row=r, column=2).value for r in range(10, 19)]
        labels = [str(x).strip() for x in labels if x is not None]
        return g_to_group, labels
    finally:
        wb.close()


class _Agg:
    """Bitta guruh (yoki jami) uchun xom yig'indilar."""

    def __init__(self) -> None:
        self.detected_c = self.detected_s = 0.0
        self.control_c = self.control_s = 0.0
        self.not_controlled_c = self.not_controlled_s = 0.0
        self.closed_c = self.closed_s = 0.0
        self.approved_c = self.approved_s = 0.0

    def add_row(self, row: dict) -> None:
        self.detected_c += _num(row.get(COL_DETECTED[0]))
        self.detected_s += _num(row.get(COL_DETECTED[1]))
        self.control_c += _num(row.get(COL_CONTROL[0]))
        self.control_s += _num(row.get(COL_CONTROL[1]))
        self.not_controlled_c += _num(row.get(COL_NOT_CONTROLLED[0]))
        self.not_controlled_s += _num(row.get(COL_NOT_CONTROLLED[1]))
        self.closed_c += _num(row.get(COL_CLOSED[0]))
        self.closed_s += _num(row.get(COL_CLOSED[1]))
        for c_col, s_col in APPROVED_COLUMN_PAIRS:
            self.approved_c += _num(row.get(c_col))
            self.approved_s += _num(row.get(s_col))

    def add_agg(self, other: "_Agg") -> None:
        self.detected_c += other.detected_c
        self.detected_s += other.detected_s
        self.control_c += other.control_c
        self.control_s += other.control_s
        self.not_controlled_c += other.not_controlled_c
        self.not_controlled_s += other.not_controlled_s
        self.closed_c += other.closed_c
        self.closed_s += other.closed_s
        self.approved_c += other.approved_c
        self.approved_s += other.approved_s


def _aggregate(rows: list[dict]) -> tuple[dict[str, _Agg], _Agg]:
    g_to_group, labels = _group_config()
    groups = {label: _Agg() for label in labels}
    total = _Agg()
    for row in rows:
        label = g_to_group.get((row.get("g_name") or "").strip())
        if label in groups:
            groups[label].add_row(row)
    for label in labels:
        total.add_agg(groups[label])
    return groups, total


def build_kpi(rows: list[dict]) -> dict:
    _, t = _aggregate(rows)
    return {
        "jami": {"count": fmt_count(t.detected_c), "sum": fmt_sum(t.detected_s)},
        "nazoratda": {
            "count": fmt_count(t.control_c),
            "sum": fmt_sum(t.control_s),
            "pct_count": fmt_pct(t.control_c, t.detected_c),
            "pct_sum": fmt_pct(t.control_s, t.detected_s),
        },
        "olinmagan": {
            "count": fmt_count(t.not_controlled_c),
            "sum": fmt_sum(t.not_controlled_s),
            "pct_count": fmt_pct(t.not_controlled_c, t.detected_c),
            "pct_sum": fmt_pct(t.not_controlled_s, t.detected_s),
        },
        "echilgan": {
            "count": fmt_count(t.closed_c),
            "sum": fmt_sum(t.closed_s),
            "pct_count": fmt_pct(t.closed_c, t.detected_c),
            "pct_sum": fmt_pct(t.closed_s, t.detected_s),
        },
        "tasdiqlangan": {
            "count": fmt_count(t.approved_c),
            "sum": fmt_sum(t.approved_s),
            # Sheet CI/CJ: тасдиғини топган / назоратдан ечилган.
            "pct_count": fmt_pct(t.approved_c, t.closed_c),
            "pct_sum": fmt_pct(t.approved_s, t.closed_s),
        },
    }


def build_cards(rows: list[dict]) -> list[dict]:
    groups, _ = _aggregate(rows)
    _, labels = _group_config()
    cards = []
    for idx, label in enumerate(labels):
        a = groups[label]
        gradient = _GREEN if idx in _GREEN_POSITIONS else _BLUE
        cards.append(
            {
                "title": label,
                "gradient": gradient,
                "employee": config.EMPLOYEE_BY_GROUP.get(label, ""),
                "detected": {"count": fmt_count(a.detected_c), "sum": fmt_sum(a.detected_s)},
                "control": {"count": fmt_count(a.control_c), "sum": fmt_sum(a.control_s)},
                # Sheet AM/AN: Назоратга олинган / Аниқланган.
                "pct": {
                    "count": fmt_pct(a.control_c, a.detected_c),
                    "sum": fmt_pct(a.control_s, a.detected_s),
                },
                # Sheet BC/BD: Назоратга олинмаган / Аниқланган.
                "not_controlled": {
                    "count": fmt_count(a.not_controlled_c),
                    "sum": fmt_sum(a.not_controlled_s),
                    "pct_count": fmt_pct(a.not_controlled_c, a.detected_c),
                    "pct_sum": fmt_pct(a.not_controlled_s, a.detected_s),
                },
                # Sheet CE/CF: Назоратдан ечилган / (Назоратга олинган + Ечилган).
                "closed": {
                    "count": fmt_count(a.closed_c),
                    "sum": fmt_sum(a.closed_s),
                    "pct_count": fmt_pct(a.closed_c, a.control_c + a.closed_c),
                    "pct_sum": fmt_pct(a.closed_s, a.control_s + a.closed_s),
                },
                # Sheet CI/CJ: Тасдиғини топган / Назоратдан ечилган.
                "approved": {
                    "count": fmt_count(a.approved_c),
                    "sum": fmt_sum(a.approved_s),
                    "pct_count": fmt_pct(a.approved_c, a.closed_c),
                    "pct_sum": fmt_pct(a.approved_s, a.closed_s),
                },
            }
        )
    return cards


def render_html(rows: list[dict]) -> str:
    env = Environment(loader=FileSystemLoader(str(config.TEMPLATES_DIR)), autoescape=True)
    template = env.get_template("daily_report.html")
    return template.render(kpi=build_kpi(rows), cards=build_cards(rows))


def build() -> Path:
    rows = run_query(config.QUERIES_DIR / "daily_summary.sql")
    html = render_html(rows)

    config.OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = config.OUTPUT_DIR / f"kunlik_hisobot_{dt.date.today().isoformat()}.pdf"

    with sync_playwright() as p:
        # --no-sandbox: konteynerda root sifatida ishlaganda chromium shusiz
        # ishga tushmaydi. Docker'da /dev/shm kichik bo'lsa ham xavfsizroq.
        browser = p.chromium.launch(args=["--no-sandbox", "--disable-dev-shm-usage"])
        try:
            page = browser.new_page()
            page.set_content(html, wait_until="networkidle")
            # Shablon CSS'da @page { size: 1908pt 1152pt } berilgan - o'sha
            # o'lchamdan foydalanamiz (aks holda kontent sig'may oppoq chiqadi).
            page.pdf(
                path=str(out_path),
                prefer_css_page_size=True,
                print_background=True,
            )
        finally:
            browser.close()

    return out_path


if __name__ == "__main__":
    print(build())
