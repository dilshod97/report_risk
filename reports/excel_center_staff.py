from __future__ import annotations

import datetime as dt
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

import config
from db import run_query

RAW_SHEET_NAME = "Xom ma'lumot"


def build_workbook(rows: list[dict]) -> Workbook:
    wb = Workbook()
    build_raw_sheet(wb, rows)
    # Namuna Excel fayli berilgach, shu yerga qo'shimcha sheetlar
    # (masalan pivot/summary) qo'shiladi: build_summary_sheet(wb, rows) kabi.
    return wb


def build_raw_sheet(wb: Workbook, rows: list[dict]) -> None:
    ws = wb.active
    ws.title = RAW_SHEET_NAME

    if not rows:
        ws.append(["Ma'lumot topilmadi"])
        return

    headers = list(rows[0].keys())
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in rows:
        ws.append([row.get(h) for h in headers])

    ws.freeze_panes = "A2"
    for idx, header in enumerate(headers, start=1):
        max_len = max(
            [len(str(header))] + [len(str(row.get(header, ""))) for row in rows]
        )
        ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 40)


def build() -> Path:
    rows = run_query(config.QUERIES_DIR / "center_staff.sql")
    wb = build_workbook(rows)

    config.OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = config.OUTPUT_DIR / f"markaz_hodimlar_{dt.date.today().isoformat()}.xlsx"
    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    print(build())
